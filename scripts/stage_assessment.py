"""Stage a raw assessment export into assessment/<run-id>/input/ + manifest.json.

The staged folder is the agentic-only folder: the agent works from manifest.json
and never sees the raw export. See docs/assessments/assessment-analysis-workflow.md.

Usage:
    python3 scripts/stage_assessment.py --export <dir> --run-id <id> [--root <dir>]
    python3 scripts/stage_assessment.py --self-test

Stdlib-only by design (mirrors stage_data.py). xlsx -> csv conversion is
best-effort: it uses openpyxl when installed and records a manifest note
otherwise, so the agent knows what it is looking at.

Exit codes: 0 staged, 1 nothing staged or error (empty input = staging at
fault, per the workflow error table).
"""

import argparse
import csv
import datetime
import hashlib
import json
import shutil
import sys
import tempfile
from pathlib import Path

STAGED_SUFFIXES = {".csv", ".json", ".md", ".txt"}
XLSX_SUFFIX = ".xlsx"
BINARY_SUFFIXES = {".bin", ".png", ".jpg", ".jpeg", ".gif", ".pdf", ".zip", ".gz", ".xls"}
MAX_MB_DEFAULT = 1.0
MAX_ROWS_DEFAULT = 1000


def _read_text(path: Path) -> tuple[str, str | None]:
    """Return (text, note) where note records an encoding conversion."""
    data = path.read_bytes()
    if b"\x00" in data[:8192]:
        raise ValueError("binary content (NUL byte)")
    try:
        return data.decode("utf-8"), None
    except UnicodeDecodeError:
        return data.decode("latin-1"), "non-utf8, converted from latin-1"


def _csv_row_count(text: str) -> int:
    try:
        return sum(1 for _ in csv.reader(text.splitlines())) - 1
    except csv.Error:
        return -1


def _json_entry_count(text: str) -> int | None:
    try:
        value = json.loads(text)
    except (json.JSONDecodeError, ValueError):
        return None
    if isinstance(value, list):
        return len(value)
    return None


def _convert_xlsx(src: Path, dst_dir: Path) -> list[tuple[str, str | None]]:
    """Convert each sheet to csv. Returns [(relative_name, note)]. Raises if
    openpyxl is unavailable or the file cannot be read."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl unavailable")
    wb = load_workbook(src, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        name = f"{src.stem}__{ws.title.lower().replace(' ', '_')}.csv"
        with open(dst_dir / name, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if c is None else c for c in row])
        out.append((name, f"converted from xlsx sheet '{ws.title}'"))
    wb.close()
    return out


def _detect_run_format(files: list[dict], input_dir: Path) -> dict:
    """Detect the assessment engine that produced the staged files.

    Returns {"run_format": <id>} plus {"run_schema_version": <value>} for tracker
    runs. Recognizes:
      - tracker run-model JSON (basename run-model*.json, legacy exports
        timestamp-suffix the filename) with a top-level `schemaVersion` key;
        the value is recorded, never pinned (legacy exports carry older values)
      - TenantSmith assess bundle (top-level `contractType ==
        "tenantsmith.assessRun.v1"`, the discriminator all TenantSmith
        contracts use)
      - anything else: "mixed"
    """
    for entry in files:
        path = input_dir / entry["name"]
        if path.suffix.lower() != ".json" or not path.stem.startswith("run-model"):
            continue
        try:
            data = json.loads(path.read_text("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError, OSError):
            continue
        if isinstance(data, dict) and "schemaVersion" in data:
            return {"run_format": "tracker-run-model",
                    "run_schema_version": str(data["schemaVersion"])}
    for entry in files:
        path = input_dir / entry["name"]
        if path.suffix.lower() != ".json":
            continue
        try:
            data = json.loads(path.read_text("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError, OSError):
            continue
        if isinstance(data, dict) and data.get("contractType") == "tenantsmith.assessRun.v1":
            return {"run_format": "tenantsmith-assess-v1"}
    return {"run_format": "mixed"}


def stage(export: Path, run_id: str, root: Path,
          max_mb: float = MAX_MB_DEFAULT, max_rows: int = MAX_ROWS_DEFAULT) -> tuple[list[dict], list[dict]]:
    """Stage export into root/run_id/input/. Returns (files, skipped)."""
    input_dir = root / run_id / "input"
    input_dir.mkdir(parents=True, exist_ok=True)
    files, skipped = [], []
    seen = {}

    for src in sorted(export.rglob("*")):
        if src.is_dir():
            continue
        if src.suffix.lower() in BINARY_SUFFIXES:
            skipped.append({"name": src.name, "reason": "binary, stripped"})
            continue
        if src.suffix.lower() not in STAGED_SUFFIXES | {XLSX_SUFFIX}:
            skipped.append({"name": src.name, "reason": f"unsupported format {src.suffix}"})
            continue

        if src.suffix.lower() == XLSX_SUFFIX:
            try:
                names = _convert_xlsx(src, input_dir)
            except (RuntimeError, Exception) as exc:
                dst = input_dir / src.name
                shutil.copy2(src, dst)
                files.append(_entry(dst, rows=None,
                                    note=f"xlsx not converted ({exc}); sheet layout must be trivial to read"))
            else:
                for name, note in names:
                    files.append(_entry(input_dir / name, rows=_csv_row_count((input_dir / name).read_text("utf-8")), note=note))
            continue

        try:
            text, enc_note = _read_text(src)
        except ValueError as exc:
            skipped.append({"name": src.name, "reason": str(exc)})
            continue

        key = (src.suffix, hashlib.sha256(text.encode("utf-8")).hexdigest())
        if key in seen:
            skipped.append({"name": src.name, "reason": f"duplicate of {seen[key]} (identical content)"})
            continue
        seen[key] = src.name

        dst = input_dir / src.name
        note = enc_note
        if dst.exists():
            dst = input_dir / f"{src.stem}_1{src.suffix}"

        if src.suffix.lower() == ".csv" and len(text.encode("utf-8")) > max_mb * 1024 * 1024:
            rows = text.splitlines()
            sampled = rows[: max_rows + 1]
            dst = input_dir / f"{src.stem}.sampled.csv"
            dst.write_text("\n".join(sampled) + "\n", encoding="utf-8")
            files.append(_entry(dst, rows=max_rows,
                                note=f"truncated: original {max(len(rows) - 1, 0)} data rows, kept {max_rows} (cap {max_mb}MB)"))
            continue

        dst.write_text(text, encoding="utf-8")
        if src.suffix.lower() == ".csv":
            files.append(_entry(dst, rows=_csv_row_count(text), note=note))
        elif src.suffix.lower() == ".json":
            files.append(_entry(dst, rows=_json_entry_count(text), note=note))
        else:
            files.append(_entry(dst, rows=len(text.splitlines()), note=note))

    if not files:
        shutil.rmtree(input_dir.parent, ignore_errors=True)
        try:
            root.rmdir()
        except OSError:
            pass
        return files, skipped
    files.sort(key=lambda f: f["name"])
    manifest = {
        "run_id": run_id,
        "staged_at": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
        "source": str(export),
        "run_format": _detect_run_format(files, input_dir),
        "files": files,
        "skipped": skipped,
    }
    (input_dir / "manifest.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    return files, skipped


def _entry(path: Path, rows: int | None, note: str | None) -> dict:
    return {
        "name": path.name,
        "format": path.suffix.lstrip(".").lower(),
        "rows": rows,
        "size": path.stat().st_size,
        "notes": [note] if note else [],
    }


def _self_test() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        export = Path(tmp) / "export"
        (export / "sub").mkdir(parents=True)
        big = ",".join(f"col{i}" for i in range(5)) + "\n" + "\n".join(
            ",".join(f"r{j}c{i}" for i in range(5)) for j in range(5000)) + "\n"
        (export / "results.csv").write_text(big, encoding="utf-8")
        (export / "small.csv").write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
        (export / "small_dup.csv").write_text("a,b\n1,2\n3,4\n", encoding="utf-8")
        (export / "latin.md").write_bytes("café\n".encode("latin-1"))
        (export / "notes.json").write_text('["one", "two", "three"]', encoding="utf-8")
        (export / "logo.png").write_bytes(b"\x89PNG\x00binary")
        (export / "sub" / "deep.txt").write_text("deep\nfile\n", encoding="utf-8")

        root = Path(tmp) / "assessment"
        files, skipped = stage(export, "run-1", root, max_mb=0.01)

        assert (root / "run-1" / "input" / "manifest.json").is_file(), "manifest missing"
        sampled = [f for f in files if f["name"] == "results.sampled.csv"]
        assert sampled and sampled[0]["rows"] == 1000, "csv cap failed"
        assert any(f["name"] == "small.csv" for f in files), "small csv missing"
        dup = [s for s in skipped if s["reason"].startswith("duplicate")]
        assert dup and dup[0]["name"] == "small_dup.csv", "dedupe failed"
        latin = [f for f in files if f["name"] == "latin.md"]
        assert latin and "non-utf8" in latin[0]["notes"][0], "encoding conversion failed"
        assert any(f["name"] == "notes.json" and f["rows"] == 3 for f in files), "json count failed"
        assert any(f["name"] == "logo.png" for s in skipped for f in [s]), "binary strip failed"
        assert any(f["name"] == "deep.txt" for f in files), "nested file missing"
        assert any(s["name"] == "logo.png" and "binary" in s["reason"] for s in skipped)

        manifest = json.loads((root / "run-1" / "input" / "manifest.json").read_text("utf-8"))
        assert manifest["run_id"] == "run-1" and len(manifest["files"]) == len(files)
        assert manifest["run_format"] == {"run_format": "mixed"}, "mixed export misdetected"

        tracker_export = Path(tmp) / "tracker-export"
        tracker_export.mkdir()
        (tracker_export / "run-model_20260101_000000.json").write_text(
            json.dumps({"schemaVersion": "2026-03-architecture-reset-v1",
                        "summary": {"controls": 42}}), encoding="utf-8")
        tf, _ = stage(tracker_export, "run-tracker", root)
        tman = json.loads((root / "run-tracker" / "input" / "manifest.json").read_text("utf-8"))
        assert tman["run_format"] == {"run_format": "tracker-run-model",
                                      "run_schema_version": "2026-03-architecture-reset-v1"}, \
            f"tracker run misdetected: {tman['run_format']}"
        assert tf[0]["name"] == "run-model_20260101_000000.json", "timestamp-suffixed run-model not staged"

        assess_export = Path(tmp) / "assess-export"
        assess_export.mkdir()
        (assess_export / "assess-bundle.json").write_text(
            json.dumps({"contractType": "tenantsmith.assessRun.v1",
                        "collectionHealth": []}), encoding="utf-8")
        af, _ = stage(assess_export, "run-assess", root)
        aman = json.loads((root / "run-assess" / "input" / "manifest.json").read_text("utf-8"))
        assert aman["run_format"] == {"run_format": "tenantsmith-assess-v1"}, \
            f"assess bundle misdetected: {aman['run_format']}"
        assert af[0]["name"] == "assess-bundle.json"

        print(f"[stage-assessment] self-test PASS ({len(files)} mixed + {len(tf)} tracker + {len(af)} assess staged)")
        return 0
    # unreachable


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--export", help="raw assessment export directory")
    parser.add_argument("--run-id", help="run identifier, e.g. acme-2026-08")
    parser.add_argument("--root", default="assessment", help="assessment root (default: ./assessment)")
    parser.add_argument("--max-mb", type=float, default=MAX_MB_DEFAULT, help="csv cap in MB (default 1.0)")
    parser.add_argument("--max-rows", type=int, default=MAX_ROWS_DEFAULT, help="sampled csv row cap (default 1000)")
    parser.add_argument("--self-test", action="store_true", help="run the runnable check and exit")
    args = parser.parse_args()

    if args.self_test:
        return _self_test()
    if not args.export or not args.run_id:
        parser.error("--export and --run-id are required")

    files, skipped = stage(Path(args.export), args.run_id, Path(args.root),
                           max_mb=args.max_mb, max_rows=args.max_rows)
    if not files:
        print("[stage-assessment] no stageable files in export; staging at fault, nothing to analyze",
              file=sys.stderr)
        return 1
    skipped_summary = f" ({len(skipped)} skipped: " + ", ".join(s["name"] for s in skipped[:5]) + ("..." if len(skipped) > 5 else ")") if skipped else ""
    print(f"[stage-assessment] staged {len(files)} files into {args.root}/{args.run_id}/input/{skipped_summary}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
